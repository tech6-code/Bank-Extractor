"""Bank Statement PDF to Excel Converter.

Usage:
    python main.py <pdf_file_or_folder> [--output <output.xlsx>]

Examples:
    python main.py statement.pdf
    python main.py statement.pdf --output result.xlsx
    python main.py ./pdf_folder/
"""

import argparse
import sys
from pathlib import Path

from extractor.pdf_parser import extract_tables_from_pdf
from extractor.excel_writer import write_tables_to_excel, write_tables_to_sheet


def get_pdf_files(input_path: str) -> list[Path]:
    """Get list of PDF files from a file path or directory."""
    path = Path(input_path)
    if path.is_file():
        if path.suffix.lower() != ".pdf":
            print(f"Error: {path} is not a PDF file.")
            sys.exit(1)
        return [path]
    elif path.is_dir():
        pdfs = sorted(path.glob("*.pdf"))
        if not pdfs:
            print(f"Error: No PDF files found in {path}")
            sys.exit(1)
        return pdfs
    else:
        print(f"Error: {input_path} does not exist.")
        sys.exit(1)


def process_single_pdf(pdf_path: Path, output_path: Path) -> None:
    """Process a single PDF and write to Excel."""
    print(f"Processing: {pdf_path.name}")
    tables = extract_tables_from_pdf(str(pdf_path))

    if not tables:
        print(f"  Warning: No data found in {pdf_path.name}")
        return

    total_rows = sum(len(t) for t in tables)
    print(f"  Found {len(tables)} table(s) with {total_rows} total row(s)")

    result = write_tables_to_excel(tables, str(output_path))
    print(f"  Output saved to: {result}")


def process_multiple_pdfs(pdf_files: list[Path], output_path: Path) -> None:
    """Process multiple PDFs into a single Excel file with separate sheets."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for pdf_path in pdf_files:
        print(f"Processing: {pdf_path.name}")
        tables = extract_tables_from_pdf(str(pdf_path))

        if not tables:
            print(f"  Warning: No data found in {pdf_path.name}")
            continue

        total_rows = sum(len(t) for t in tables)
        print(f"  Found {len(tables)} table(s) with {total_rows} total row(s)")

        # Create sheet with PDF filename (truncated to 31 chars for Excel limit)
        ws = wb.create_sheet(title=pdf_path.stem[:31])
        write_tables_to_sheet(ws, tables)

    if not wb.sheetnames:
        print("Error: No data extracted from any PDF files.")
        sys.exit(1)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\nOutput saved to: {output_path.resolve()}")


def main():
    parser = argparse.ArgumentParser(
        description="Convert bank statement PDFs to Excel files."
    )
    parser.add_argument(
        "input",
        help="Path to a PDF file or a folder containing PDF files.",
    )
    parser.add_argument(
        "--output", "-o",
        help="Output Excel file path (default: <input_name>.xlsx)",
    )
    args = parser.parse_args()

    pdf_files = get_pdf_files(args.input)

    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        input_path = Path(args.input)
        if input_path.is_file():
            output_path = input_path.with_suffix(".xlsx")
        else:
            output_path = input_path / "bank_statements.xlsx"

    print(f"Bank Statement PDF to Excel Converter")
    print(f"{'=' * 40}")

    if len(pdf_files) == 1:
        process_single_pdf(pdf_files[0], output_path)
    else:
        print(f"Found {len(pdf_files)} PDF file(s)")
        process_multiple_pdfs(pdf_files, output_path)

    print("\nDone!")


if __name__ == "__main__":
    main()
