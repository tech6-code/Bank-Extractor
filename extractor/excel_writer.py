from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_BODY_FONT = Font(size=11)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_tables_to_sheet(ws, tables: list[list[list[str]]]) -> None:
    """Write tables into an existing worksheet with formatting.

    The first row of each table is styled as a header. A blank row is
    inserted between consecutive tables.
    """
    current_row = 1
    for table in tables:
        if not table:
            continue
        for row_idx, row_data in enumerate(table):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=current_row, column=col_idx + 1, value=value)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if row_idx == 0:
                    cell.font = _HEADER_FONT
                    cell.fill = _HEADER_FILL
                else:
                    cell.font = _BODY_FONT
            current_row += 1
        current_row += 1  # blank row between tables

    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col if cell.value), default=0
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)


def write_tables_to_excel(
    tables: list[list[list[str]]],
    output_path: str,
    sheet_name: str = "Bank Statement",
) -> str:
    """Write extracted tables to a new Excel file.

    Args:
        tables: List of tables (each table is a list of rows).
        output_path: Path for the output Excel file.
        sheet_name: Name of the worksheet.

    Returns:
        The absolute path of the created Excel file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    write_tables_to_sheet(ws, tables)

    wb.save(output_path)
    return str(output_path.resolve())
