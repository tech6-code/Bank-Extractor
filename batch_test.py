"""
Batch test all PDFs in a directory — extract, validate, detect bank, save templates.

Usage:
    python batch_test.py [directory]

Default directory: C:\\Users\\GC-IT-4\\Documents\\Bank Statement Module\\Digital

Output:
    - Live progress printed to console
    - Full summary table at the end
    - batch_test_results.csv   (always)
    - batch_test_results.xlsx  (if openpyxl available)
"""

import sys
import os
import time
import traceback
import csv
from pathlib import Path
from datetime import datetime

# ---------------------------------------------------------------------------
# Bootstrap: make sure the project root is on sys.path
# ---------------------------------------------------------------------------
ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

from extractor.pdf_parser import extract_transactions
from extractor.balance_validator import validate_balances
from extractor.template_engine import match_template, save_extraction_template, detect_bank_name
import pdfplumber


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _count_pages(pdf_path: str) -> int:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            return len(pdf.pages)
    except Exception:
        try:
            import fitz
            doc = fitz.open(pdf_path)
            n = len(doc)
            doc.close()
            return n
        except Exception:
            return -1


def _separate_meta(raw: list) -> tuple[list, dict]:
    """Split metadata entry from transaction list (mirrors server.py logic)."""
    txns = [t for t in raw if "_extraction_meta" not in t]
    metas = [t for t in raw if "_extraction_meta" in t]
    meta = metas[0]["_extraction_meta"] if metas else {}
    return txns, meta


def _total_amount(txns: list, field: str) -> float:
    total = 0.0
    for t in txns:
        v = t.get(field, "")
        if v:
            try:
                total += abs(float(str(v).replace(",", "")))
            except ValueError:
                pass
    return total


# ---------------------------------------------------------------------------
# Per-PDF test
# ---------------------------------------------------------------------------

def test_pdf(pdf_path: str) -> dict:
    result = {
        "file":             Path(pdf_path).name,
        "pages":            0,
        "bank":             "",
        "transactions":     0,
        "strategy":         "N/A",
        "template_status":  "N/A",   # matched / new-saved / not-saved / error
        "template_id":      "",
        "debits":           0,
        "credits":          0,
        "has_balance":      0,
        "total_debit":      0.0,
        "total_credit":     0.0,
        "validated_rows":   0,
        "accuracy_pct":     None,
        "mismatches":       0,
        "direction":        "",
        "error":            "",
        "time_sec":         0.0,
    }

    start = time.time()
    try:
        # 1 ── Page count
        result["pages"] = _count_pages(pdf_path)

        # 2 ── Template matching
        template = None
        try:
            template = match_template(pdf_path)
        except Exception as e:
            pass  # non-fatal

        if template:
            result["template_status"] = "matched"
            result["template_id"]     = str(template.get("id", ""))

        # 3 ── Bank name (parallel, fast)
        try:
            bank = detect_bank_name(pdf_path)
            result["bank"] = bank or ""
        except Exception:
            pass

        # 4 ── Extraction
        raw = extract_transactions(pdf_path, template=template)
        txns, meta = _separate_meta(raw)

        result["transactions"] = len(txns)
        result["strategy"]     = meta.get("strategy", "unknown")

        if not txns:
            result["error"] = "No transactions extracted"
            result["time_sec"] = round(time.time() - start, 1)
            return result

        # 5 ── Save template if it's a new layout
        if not template and meta:
            try:
                col_map   = meta.get("col_map", {})
                strategy  = meta.get("strategy", "unknown")
                tid = save_extraction_template(
                    pdf_path=pdf_path,
                    col_map=col_map,
                    strategy=strategy,
                    transactions_count=len(txns),
                )
                if tid:
                    result["template_status"] = "new-saved"
                    result["template_id"]     = str(tid)
                else:
                    result["template_status"] = "not-saved"
            except Exception as e:
                result["template_status"] = f"save-err"

        # 6 ── Amount totals
        result["debits"]       = sum(1 for t in txns if t.get("debit"))
        result["credits"]      = sum(1 for t in txns if t.get("credit"))
        result["has_balance"]  = sum(1 for t in txns if t.get("balance"))
        result["total_debit"]  = _total_amount(txns, "debit")
        result["total_credit"] = _total_amount(txns, "credit")

        # 7 ── Balance validation
        v = validate_balances(txns)
        result["validated_rows"] = v.get("validated_rows", 0)
        result["accuracy_pct"]   = v.get("accuracy_pct")
        result["mismatches"]     = v.get("invalid_rows", 0)
        result["direction"]      = v.get("direction", "")

    except Exception as e:
        result["error"] = str(e)[:120]
        traceback.print_exc()

    result["time_sec"] = round(time.time() - start, 1)
    return result


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

COL_WIDTHS = {
    "#":             4,
    "File":          42,
    "Bank":          18,
    "Pg":             3,
    "Txns":           5,
    "Strategy":      11,
    "Template":      11,
    "TID":            5,
    "Dr":             5,
    "Cr":             5,
    "Bal":            4,
    "TotalDr":       14,
    "TotalCr":       14,
    "ValRows":        7,
    "Accuracy":       9,
    "Mis":            4,
    "Dir":           11,
    "Time":           5,
    "Status":         9,
}


def _hdr() -> str:
    return (
        f"{'#':>{COL_WIDTHS['#']}} "
        f"{'File':<{COL_WIDTHS['File']}} "
        f"{'Bank':<{COL_WIDTHS['Bank']}} "
        f"{'Pg':>{COL_WIDTHS['Pg']}} "
        f"{'Txns':>{COL_WIDTHS['Txns']}} "
        f"{'Strategy':<{COL_WIDTHS['Strategy']}} "
        f"{'Template':<{COL_WIDTHS['Template']}} "
        f"{'TID':>{COL_WIDTHS['TID']}} "
        f"{'Dr':>{COL_WIDTHS['Dr']}} "
        f"{'Cr':>{COL_WIDTHS['Cr']}} "
        f"{'Bal':>{COL_WIDTHS['Bal']}} "
        f"{'TotalDebit':>{COL_WIDTHS['TotalDr']}} "
        f"{'TotalCredit':>{COL_WIDTHS['TotalCr']}} "
        f"{'ValRows':>{COL_WIDTHS['ValRows']}} "
        f"{'Accuracy':>{COL_WIDTHS['Accuracy']}} "
        f"{'Mis':>{COL_WIDTHS['Mis']}} "
        f"{'Direction':<{COL_WIDTHS['Dir']}} "
        f"{'Time':>{COL_WIDTHS['Time']}} "
        f"Status"
    )


def _row(i: int, r: dict) -> str:
    acc = f"{r['accuracy_pct']:.1f}%" if r["accuracy_pct"] is not None else "-"
    status = (
        "ERROR"   if r["error"] and r["transactions"] == 0 else
        "NO DATA" if r["transactions"] == 0 else
        "WARN"    if r["error"] else
        "OK"
    )
    return (
        f"{i:>{COL_WIDTHS['#']}} "
        f"{r['file'][:COL_WIDTHS['File']]:<{COL_WIDTHS['File']}} "
        f"{r['bank'][:COL_WIDTHS['Bank']]:<{COL_WIDTHS['Bank']}} "
        f"{r['pages']:>{COL_WIDTHS['Pg']}} "
        f"{r['transactions']:>{COL_WIDTHS['Txns']}} "
        f"{r['strategy'][:COL_WIDTHS['Strategy']]:<{COL_WIDTHS['Strategy']}} "
        f"{r['template_status'][:COL_WIDTHS['Template']]:<{COL_WIDTHS['Template']}} "
        f"{r['template_id'][:COL_WIDTHS['TID']]:>{COL_WIDTHS['TID']}} "
        f"{r['debits']:>{COL_WIDTHS['Dr']}} "
        f"{r['credits']:>{COL_WIDTHS['Cr']}} "
        f"{r['has_balance']:>{COL_WIDTHS['Bal']}} "
        f"{r['total_debit']:>{COL_WIDTHS['TotalDr']},.2f} "
        f"{r['total_credit']:>{COL_WIDTHS['TotalCr']},.2f} "
        f"{r['validated_rows']:>{COL_WIDTHS['ValRows']}} "
        f"{acc:>{COL_WIDTHS['Accuracy']}} "
        f"{r['mismatches']:>{COL_WIDTHS['Mis']}} "
        f"{r['direction'][:COL_WIDTHS['Dir']]:<{COL_WIDTHS['Dir']}} "
        f"{r['time_sec']:>{COL_WIDTHS['Time']}.1f}s "
        f"{status}"
    )


def save_csv(results: list[dict], path: Path):
    fields = [
        "file", "bank", "pages", "transactions", "strategy",
        "template_status", "template_id",
        "debits", "credits", "has_balance",
        "total_debit", "total_credit",
        "validated_rows", "accuracy_pct", "mismatches", "direction",
        "time_sec", "error",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(results)
    print(f"\nCSV saved: {path}")


def save_excel(results: list[dict], path: Path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not available — skipping Excel export")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Test Results"

    headers = [
        "#", "File", "Bank", "Pages", "Transactions", "Strategy",
        "Template Status", "Template ID",
        "Debit Entries", "Credit Entries", "Balance Entries",
        "Total Debit", "Total Credit",
        "Validated Rows", "Accuracy %", "Mismatches", "Direction",
        "Time (s)", "Error",
    ]

    # Header row styling
    hdr_fill   = PatternFill("solid", fgColor="1F3864")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    thin_side  = Side(style="thin", color="CCCCCC")
    thin_border= Border(left=thin_side, right=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Fill colours for status
    ok_fill      = PatternFill("solid", fgColor="E8F5E9")
    warn_fill    = PatternFill("solid", fgColor="FFF9C4")
    error_fill   = PatternFill("solid", fgColor="FFEBEE")
    nodata_fill  = PatternFill("solid", fgColor="F3E5F5")
    perfect_fill = PatternFill("solid", fgColor="C8E6C9")

    for row_idx, r in enumerate(results, 2):
        status = (
            "ERROR"   if r["error"] and r["transactions"] == 0 else
            "NO DATA" if r["transactions"] == 0 else
            "OK"
        )
        row_fill = (
            error_fill  if status == "ERROR"   else
            nodata_fill if status == "NO DATA" else
            ok_fill
        )
        acc = r.get("accuracy_pct")
        if acc == 100.0:
            row_fill = perfect_fill

        values = [
            row_idx - 1,
            r["file"],
            r["bank"],
            r["pages"],
            r["transactions"],
            r["strategy"],
            r["template_status"],
            r["template_id"] or "",
            r["debits"],
            r["credits"],
            r["has_balance"],
            round(r["total_debit"], 2),
            round(r["total_credit"], 2),
            r["validated_rows"],
            acc,
            r["mismatches"],
            r["direction"],
            r["time_sec"],
            r["error"] or "",
        ]
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = row_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal="left")
            if col_idx in (12, 13):  # Total debit/credit
                c.number_format = "#,##0.00"
            if col_idx == 15 and val is not None:  # Accuracy %
                c.number_format = "0.0"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total   = len(results)
    ok_cnt  = sum(1 for r in results if not r["error"] and r["transactions"] > 0)
    err_cnt = sum(1 for r in results if r["error"] and r["transactions"] == 0)
    nd_cnt  = sum(1 for r in results if not r["error"] and r["transactions"] == 0)
    total_txns = sum(r["transactions"] for r in results)
    total_debit= sum(r["total_debit"] for r in results)
    total_credit=sum(r["total_credit"] for r in results)
    with_acc   = [r["accuracy_pct"] for r in results if r["accuracy_pct"] is not None]
    avg_acc    = sum(with_acc) / len(with_acc) if with_acc else None
    perfect    = sum(1 for a in with_acc if a == 100.0)
    strats     = {}
    for r in results:
        strats[r["strategy"]] = strats.get(r["strategy"], 0) + 1
    tmpl_new     = sum(1 for r in results if r["template_status"] == "new-saved")
    tmpl_matched = sum(1 for r in results if r["template_status"] == "matched")

    summary_rows = [
        ("Run date",                datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total PDFs",              total),
        ("Successful extractions",  ok_cnt),
        ("No data",                 nd_cnt),
        ("Errors",                  err_cnt),
        ("",                        ""),
        ("Total transactions",      total_txns),
        ("Total debit (sum)",       round(total_debit, 2)),
        ("Total credit (sum)",      round(total_credit, 2)),
        ("",                        ""),
        ("PDFs with balance data",  len(with_acc)),
        ("Average accuracy %",      round(avg_acc, 2) if avg_acc is not None else "N/A"),
        ("100% accurate PDFs",      perfect),
        ("",                        ""),
        ("Templates matched",       tmpl_matched),
        ("New templates saved",     tmpl_new),
        ("",                        ""),
        ("Strategies used",         ""),
    ]
    for strat, cnt in sorted(strats.items(), key=lambda x: -x[1]):
        summary_rows.append((f"  {strat}", cnt))

    lbl_font = Font(bold=True, size=11)
    for r_idx, (label, value) in enumerate(summary_rows, 1):
        c1 = ws2.cell(row=r_idx, column=1, value=label)
        c2 = ws2.cell(row=r_idx, column=2, value=value)
        if label and not label.startswith(" "):
            c1.font = lbl_font
        if isinstance(value, float):
            c2.number_format = "#,##0.00"

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    # Column widths for main sheet
    col_widths = [4, 45, 20, 6, 12, 13, 13, 8, 8, 8, 8, 14, 14, 10, 10, 6, 13, 7, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    print(f"Excel saved: {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    directory = (
        sys.argv[1]
        if len(sys.argv) > 1
        else r"C:\Users\GC-IT-4\Documents\Bank Statement Module\Digital"
    )

    pdf_files = sorted(Path(directory).glob("*.[pP][dD][fF]"))
    total = len(pdf_files)
    if total == 0:
        print(f"No PDF files found in: {directory}")
        sys.exit(1)

    print("=" * 160)
    print(f"  BATCH TEST  |  {total} PDFs  |  {directory}")
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 160)
    print(_hdr())
    print("-" * 160)

    results = []
    sep_line = "-" * 160

    # Incremental CSV so partial results survive if the run is interrupted.
    incremental_csv = ROOT / f"batch_test_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.partial.csv"
    inc_fields = [
        "file", "bank", "pages", "transactions", "strategy",
        "template_status", "template_id",
        "debits", "credits", "has_balance",
        "total_debit", "total_credit",
        "validated_rows", "accuracy_pct", "mismatches", "direction",
        "time_sec", "error",
    ]
    inc_fh = open(incremental_csv, "w", newline="", encoding="utf-8")
    inc_writer = csv.DictWriter(inc_fh, fieldnames=inc_fields, extrasaction="ignore")
    inc_writer.writeheader()
    inc_fh.flush()

    for i, pdf_path in enumerate(pdf_files, 1):
        # Brief live status before processing
        print(f"  ... [{i:3d}/{total}] {pdf_path.name[:60]}", end="\r", flush=True)

        r = test_pdf(str(pdf_path))
        results.append(r)

        # Append to incremental CSV and flush
        inc_writer.writerow(r)
        inc_fh.flush()

        # Print the result row
        print(_row(i, r))

        # Every 10 files print a light separator for readability
        if i % 10 == 0 and i < total:
            print(sep_line[:80])

    inc_fh.close()

    print("=" * 160)

    # ── Final summary ────────────────────────────────────────────────────────
    total_txns   = sum(r["transactions"] for r in results)
    ok_cnt       = sum(1 for r in results if not r["error"] and r["transactions"] > 0)
    err_cnt      = sum(1 for r in results if r["error"] and r["transactions"] == 0)
    nd_cnt       = sum(1 for r in results if not r["error"] and r["transactions"] == 0)
    with_acc     = [r["accuracy_pct"] for r in results if r["accuracy_pct"] is not None]
    avg_acc      = sum(with_acc) / len(with_acc) if with_acc else None
    perfect      = sum(1 for a in with_acc if a == 100.0)
    total_debit  = sum(r["total_debit"]  for r in results)
    total_credit = sum(r["total_credit"] for r in results)
    tmpl_new     = sum(1 for r in results if r["template_status"] == "new-saved")
    tmpl_matched = sum(1 for r in results if r["template_status"] == "matched")
    strats: dict[str, int] = {}
    for r in results:
        strats[r["strategy"]] = strats.get(r["strategy"], 0) + 1
    total_time   = sum(r["time_sec"] for r in results)

    print(f"""
SUMMARY
  PDFs processed      : {total}
  Successful          : {ok_cnt}
  No data extracted   : {nd_cnt}
  Errors              : {err_cnt}
  Total transactions  : {total_txns:,}
  Total debit  (sum)  : {total_debit:>18,.2f}
  Total credit (sum)  : {total_credit:>18,.2f}
  PDFs with bal. data : {len(with_acc)}
  Average accuracy    : {f'{avg_acc:.2f}%' if avg_acc is not None else 'N/A'}
  100% accurate       : {perfect}
  Templates matched   : {tmpl_matched}
  New templates saved : {tmpl_new}
  Strategies          : {dict(sorted(strats.items(), key=lambda x:-x[1]))}
  Total time          : {total_time:.1f}s  ({total_time/60:.1f} min)
""")

    # Highlight any errors or no-data PDFs
    problem_files = [r for r in results if r["error"] or r["transactions"] == 0]
    if problem_files:
        print("PROBLEM FILES:")
        for r in problem_files:
            tag = "ERROR" if r["error"] else "NO DATA"
            msg = r["error"] or "0 transactions"
            print(f"  [{tag}] {r['file']}  —  {msg}")
        print()

    # ── Save results ─────────────────────────────────────────────────────────
    out_dir = ROOT
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path  = out_dir / f"batch_test_results_{ts}.csv"
    xlsx_path = out_dir / f"batch_test_results_{ts}.xlsx"

    save_csv(results, csv_path)
    save_excel(results, xlsx_path)

    print(f"\nDone. {total} PDFs processed in {total_time:.1f}s.")


if __name__ == "__main__":
    main()
